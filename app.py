import os
import io
import json
import uuid
import hashlib
from datetime import datetime
from functools import wraps
from flask import Flask, request, jsonify, send_file, render_template, send_from_directory, session, redirect, url_for
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'ternary-claim-secret-2026-change-in-prod')
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB

COMPANY_NAME = "Ternary Fund Management Pte Ltd"
COMPANY_UEN = "UEN: 201902851Z"
COMPANY_ADDRESS = "50 Armenian Street #02-04 Wilmer Place, Singapore 179938"

DB_PATH       = os.path.join(os.path.dirname(__file__), 'submissions.json')
USERS_PATH    = os.path.join(os.path.dirname(__file__), 'users.json')
DRAFTS_PATH   = os.path.join(os.path.dirname(__file__), 'drafts.json')
SETTINGS_PATH = os.path.join(os.path.dirname(__file__), 'settings.json')

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)


# ── Auth helpers ──────────────────────────────────────
def _hash(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

def _load_users():
    if not os.path.exists(USERS_PATH):
        defaults = [
            {'username': 'admin',    'password': _hash('TerClaim16!'),   'role': 'admin',    'display_name': 'Admin'},
            {'username': 'xingye',   'password': _hash('pass1234'),   'role': 'employee', 'display_name': 'Xingye, Zhou'},
            {'username': 'peter',    'password': _hash('pass1234'),   'role': 'employee', 'display_name': 'Peter Tan'},
            {'username': 'jason',    'password': _hash('pass1234'),   'role': 'employee', 'display_name': 'Jason Chan'},
            {'username': 'mary',     'password': _hash('pass1234'),   'role': 'employee', 'display_name': 'Mary'},
            {'username': 'chuiwhei', 'password': _hash('pass1234'),   'role': 'employee', 'display_name': 'Chui Whei'},
            {'username': 'egan',     'password': _hash('pass1234'),   'role': 'employee', 'display_name': 'Egan'},
            {'username': 'edward',   'password': _hash('pass1234'),   'role': 'employee', 'display_name': 'Edward'},
            {'username': 'thomas',   'password': _hash('pass1234'),   'role': 'employee', 'display_name': 'Thomas'},
            {'username': 'yongchuan','password': _hash('pass1234'),   'role': 'employee', 'display_name': 'Yong Chuan'},
            {'username': 'gabriel',  'password': _hash('pass1234'),   'role': 'employee', 'display_name': 'Gabriel'},
            {'username': 'hannah',   'password': _hash('hannah2026'), 'role': 'employee', 'display_name': 'Hannah'},
        ]
        with open(USERS_PATH, 'w', encoding='utf-8') as f:
            json.dump(defaults, f, indent=2)
        return defaults
    with open(USERS_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

def _get_user(username):
    return next((u for u in _load_users() if u['username'] == username), None)

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login_page'))
        if session.get('role') != 'admin':
            return jsonify({'error': 'Admin only'}), 403
        return f(*args, **kwargs)
    return decorated


# ── Submission store ──────────────────────────────────
def _load_submissions():
    if not os.path.exists(DB_PATH):
        return []
    with open(DB_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

def _save_submissions(subs):
    with open(DB_PATH, 'w', encoding='utf-8') as f:
        json.dump(subs, f, indent=2, ensure_ascii=False)


# ── Draft store ───────────────────────────────────────
def _load_drafts():
    if not os.path.exists(DRAFTS_PATH):
        return []
    with open(DRAFTS_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

def _save_drafts(drafts):
    with open(DRAFTS_PATH, 'w', encoding='utf-8') as f:
        json.dump(drafts, f, indent=2, ensure_ascii=False)


# ── Settings store ────────────────────────────────────
def _load_settings():
    if not os.path.exists(SETTINGS_PATH):
        return {'claim_no_next': 1}
    with open(SETTINGS_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

def _save_settings(settings):
    with open(SETTINGS_PATH, 'w', encoding='utf-8') as f:
        json.dump(settings, f, indent=2)

def _next_claim_no():
    """Return next claim number and increment the counter."""
    s = _load_settings()
    n = s.get('claim_no_next', 1)
    s['claim_no_next'] = n + 1
    _save_settings(s)
    return n


# ── Auth routes ───────────────────────────────────────
@app.route('/login', methods=['GET'])
def login_page():
    if 'username' in session:
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login_post():
    data = request.get_json(force=True)
    user = _get_user(data.get('username', '').strip().lower())
    if not user or user['password'] != _hash(data.get('password', '')):
        return jsonify({'error': 'Invalid username or password'}), 401
    session['username']     = user['username']
    session['role']         = user['role']
    session['display_name'] = user['display_name']
    return jsonify({'role': user['role']})

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

@app.route('/api/me')
@login_required
def me():
    return jsonify({
        'username':     session['username'],
        'role':         session['role'],
        'display_name': session['display_name'],
    })


# ── Routes ────────────────────────────────────────────
@app.route('/')
@login_required
def index():
    return render_template('index.html')


@app.route('/admin')
@login_required
def admin():
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    return render_template('admin.html')


@app.route('/settings')
@login_required
def settings_page():
    if session.get('role') != 'admin':
        return redirect(url_for('index'))
    return render_template('settings.html')


@app.route('/uploads/<path:filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


@app.route('/api/set-print-queue', methods=['POST'])
@login_required
def set_print_queue():
    data = request.get_json(force=True)
    session['print_queue'] = data.get('files', [])
    session['print_name']  = data.get('name', '')
    return jsonify({'ok': True})


@app.route('/print-receipts')
@login_required
def print_receipts_page():
    files = session.get('print_queue', [])
    name  = session.get('print_name', '')
    return render_template('print_receipts.html', files=files, name=name)


def _find_libreoffice():
    import shutil
    cmd = shutil.which('libreoffice')
    if cmd:
        return cmd
    cmd = shutil.which('soffice')
    if cmd:
        return cmd
    if os.name == 'nt':
        paths = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
        for p in paths:
            if os.path.exists(p):
                return p
    return None


def _convert_to_pdf_win32(src_path, dest_pdf_path):
    """Convert DOCX/DOC/MSG to PDF using Office COM automation on Windows."""
    import win32com.client
    import pythoncom

    pythoncom.CoInitialize()
    ext = os.path.splitext(src_path)[1].lower()

    if ext in ('.docx', '.doc'):
        word = None
        try:
            word = win32com.client.DispatchEx("Word.Application")
            word.Visible = False
            word.DisplayAlerts = 0
            doc = word.Documents.Open(os.path.abspath(src_path), ReadOnly=True)
            doc.SaveAs(os.path.abspath(dest_pdf_path), FileFormat=17) # 17 is wdFormatPDF
            doc.Close()
            return True
        finally:
            if word:
                word.Quit()
    elif ext == '.msg':
        outlook = None
        word = None
        try:
            outlook = win32com.client.DispatchEx("Outlook.Application")
            msg = outlook.CreateItemFromTemplate(os.path.abspath(src_path))
            
            temp_dir = os.path.dirname(dest_pdf_path)
            temp_html_path = os.path.join(temp_dir, "temp_msg.html")
            msg.SaveAs(os.path.abspath(temp_html_path), 5) # 5 is olHTML
            
            word = win32com.client.DispatchEx("Word.Application")
            word.Visible = False
            word.DisplayAlerts = 0
            doc = word.Documents.Open(os.path.abspath(temp_html_path), ReadOnly=True)
            doc.SaveAs(os.path.abspath(dest_pdf_path), FileFormat=17)
            doc.Close()
            
            try:
                os.remove(temp_html_path)
            except Exception:
                pass
            return True
        finally:
            if word:
                try:
                    word.Quit()
                except Exception:
                    pass
            if outlook:
                try:
                    outlook.Quit()
                except Exception:
                    pass
    return False


def _make_pdf_response(pdf_bytes, filename):
    from flask import make_response as _mr
    resp = _mr(pdf_bytes)
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = (
        'inline; filename="' + os.path.splitext(filename)[0] + '.pdf"'
    )
    return resp


@app.route('/api/to-pdf/<path:filename>')
@login_required
def convert_to_pdf(filename):
    """Convert a DOCX/DOC/MSG file in the uploads folder to PDF."""
    import subprocess, tempfile, shutil as _shutil
    src_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(src_path):
        return jsonify({'error': 'File not found'}), 404

    ext = os.path.splitext(filename)[1].lower()
    if ext == '.pdf':
        # Already a PDF — just serve it directly
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename,
                                   mimetype='application/pdf')

    # Try using LibreOffice first
    libreoffice_bin = _find_libreoffice()
    if libreoffice_bin:
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                result = subprocess.run(
                    [libreoffice_bin, '--headless', '--convert-to', 'pdf',
                     '--outdir', tmpdir, src_path],
                    capture_output=True, timeout=60
                )
                pdfs = [f for f in os.listdir(tmpdir) if f.lower().endswith('.pdf')]
                if result.returncode == 0 and pdfs:
                    pdf_path = os.path.join(tmpdir, pdfs[0])
                    with open(pdf_path, 'rb') as f:
                        pdf_bytes = f.read()
                    return _make_pdf_response(pdf_bytes, filename)
        except Exception as e:
            app.logger.warning('LibreOffice conversion failed, trying COM: %s', e)

    # Fallback to Windows COM automation (if on Windows and Office is installed)
    if os.name == 'nt':
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                dest_pdf_path = os.path.join(tmpdir, 'converted.pdf')
                if _convert_to_pdf_win32(src_path, dest_pdf_path):
                    with open(dest_pdf_path, 'rb') as f:
                        pdf_bytes = f.read()
                    return _make_pdf_response(pdf_bytes, filename)
        except Exception as e:
            app.logger.error('Windows COM conversion failed: %s', e)
            return jsonify({'error': f'Office COM conversion failed: {str(e)}'}), 500

    return jsonify({
        'error': 'No conversion tool available. Please install LibreOffice to support document conversion.'
    }), 500


@app.route('/api/upload', methods=['POST'])
@login_required
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'No filename'}), 400

    original_ext = os.path.splitext(file.filename)[1].lower()
    allowed = {'.jpg', '.jpeg', '.png', '.gif', '.pdf', '.webp', '.heic', '.msg', '.docx', '.doc'}
    if original_ext not in allowed:
        return jsonify({'error': f'File type {original_ext} not allowed'}), 400

    temp_unique_name = f"{uuid.uuid4().hex}{original_ext}"
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], temp_unique_name)
    file.save(save_path)

    # Convert Word / MSG attachments to PDF immediately on upload
    if original_ext in ('.docx', '.doc', '.msg'):
        pdf_unique_name = f"{uuid.uuid4().hex}.pdf"
        pdf_save_path = os.path.join(app.config['UPLOAD_FOLDER'], pdf_unique_name)
        
        conversion_success = False
        error_msg = ""
        
        # 1. Try LibreOffice
        libreoffice_bin = _find_libreoffice()
        if libreoffice_bin:
            try:
                import tempfile, subprocess
                with tempfile.TemporaryDirectory() as tmpdir:
                    result = subprocess.run(
                        [libreoffice_bin, '--headless', '--convert-to', 'pdf',
                         '--outdir', tmpdir, save_path],
                        capture_output=True, timeout=60
                    )
                    pdfs = [f for f in os.listdir(tmpdir) if f.lower().endswith('.pdf')]
                    if result.returncode == 0 and pdfs:
                        import shutil
                        shutil.copy(os.path.join(tmpdir, pdfs[0]), pdf_save_path)
                        conversion_success = True
            except Exception as e:
                error_msg = str(e)
                app.logger.warning('LibreOffice upload conversion failed: %s', e)

        # 2. Try Windows COM Fallback
        if not conversion_success and os.name == 'nt':
            try:
                if _convert_to_pdf_win32(save_path, pdf_save_path):
                    conversion_success = True
            except Exception as e:
                error_msg = str(e)
                app.logger.error('COM upload conversion failed: %s', e)

        # Clean up the original uploaded Word/MSG file
        try:
            os.remove(save_path)
        except Exception:
            pass

        if not conversion_success:
            return jsonify({
                'error': f'Failed to convert uploaded document to PDF: {error_msg or "No conversion tools available"}'
            }), 500

        return jsonify({
            'filename': pdf_unique_name,
            'original_name': file.filename,
            'url': f'/uploads/{pdf_unique_name}'
        })

    return jsonify({
        'filename': temp_unique_name,
        'original_name': file.filename,
        'url': f'/uploads/{temp_unique_name}'
    })


@app.route('/api/submit', methods=['POST'])
@login_required
def submit_claim():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400

    subs = _load_submissions()
    submission_id = uuid.uuid4().hex[:10]
    total = sum(float(it.get('total') or 0) for it in data.get('items', []) if it.get('total'))

    # Auto-assign claim number: use provided value or fetch+increment counter
    claim_no_auto = data.get('claim_no_auto', False)
    provided_no = str(data.get('claim_no', '')).strip()
    if claim_no_auto or not provided_no:
        claim_no = str(_next_claim_no())
    else:
        s = _load_settings()
        if provided_no == str(s.get('claim_no_next', 1)):
            _next_claim_no()   # consume this pre-filled number
        claim_no = provided_no

    record = {
        'id': submission_id,
        'submitted_at': datetime.now().isoformat(timespec='seconds'),
        'submitted_by': session['username'],
        'status': 'Pending',
        'employee_name': data.get('employee_name', ''),
        'claim_no': claim_no,
        'period_from': data.get('period_from', ''),
        'period_to': data.get('period_to', ''),
        'total': round(total, 2),
        'currency': data.get('currency', 'SGD'),
        'notes': data.get('notes', ''),
        'items': data.get('items', []),
        'attachments': data.get('attachments', []),
    }
    subs.append(record)
    _save_submissions(subs)
    return jsonify({'id': submission_id, 'claim_no': claim_no})


@app.route('/api/submissions', methods=['GET'])
@login_required
def list_submissions():
    subs = _load_submissions()
    if session.get('role') != 'admin':
        subs = [s for s in subs if s.get('submitted_by') == session['username']]
    return jsonify(subs)


@app.route('/api/submissions/<sid>', methods=['GET'])
@login_required
def get_submission(sid):
    subs = _load_submissions()
    rec = next((s for s in subs if s['id'] == sid), None)
    if not rec:
        return jsonify({'error': 'Not found'}), 404
    if session.get('role') != 'admin' and rec.get('submitted_by') != session['username']:
        return jsonify({'error': 'Forbidden'}), 403
    return jsonify(rec)


@app.route('/api/submissions/<sid>', methods=['PUT'])
@admin_required
def update_submission(sid):
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400
    subs = _load_submissions()
    rec = next((s for s in subs if s['id'] == sid), None)
    if not rec:
        return jsonify({'error': 'Not found'}), 404
    total = sum(float(it.get('total') or 0) for it in data.get('items', []) if it.get('total'))
    rec.update({
        'employee_name': data.get('employee_name', rec['employee_name']),
        'claim_no':      data.get('claim_no',      rec['claim_no']),
        'period_from':   data.get('period_from',   rec['period_from']),
        'period_to':     data.get('period_to',     rec['period_to']),
        'total':         round(total, 2),
        'currency':      data.get('currency',      rec.get('currency', 'SGD')),
        'notes':         data.get('notes',         rec.get('notes', '')),
        'items':         data.get('items',         rec['items']),
        'attachments':   data.get('attachments',   rec.get('attachments', [])),
        'last_edited_at': datetime.now().isoformat(timespec='seconds'),
        'last_edited_by': session['username'],
    })
    _save_submissions(subs)
    return jsonify({'id': sid, 'claim_no': rec['claim_no']})


@app.route('/api/submissions/<sid>', methods=['DELETE'])
@login_required
def delete_submission(sid):
    subs = _load_submissions()
    rec = next((s for s in subs if s['id'] == sid), None)
    if not rec:
        return jsonify({'error': 'Not found'}), 404

    is_admin = session.get('role') == 'admin'
    if not is_admin and rec.get('submitted_by') != session['username']:
        return jsonify({'error': 'Forbidden'}), 403

    if rec.get('status') != 'Pending':
        return jsonify({'error': 'Only pending claims can be deleted'}), 400

    updated_subs = [s for s in subs if s['id'] != sid]
    _save_submissions(updated_subs)
    return jsonify({'ok': True})


@app.route('/api/users', methods=['GET'])
@admin_required
def list_users():
    users = _load_users()
    return jsonify([{'username': u['username'], 'display_name': u['display_name']} for u in users])


@app.route('/api/submissions/<sid>/status', methods=['PATCH'])
@admin_required
def update_status(sid):
    new_status = request.get_json(force=True).get('status', '')
    if new_status not in ('Pending', 'Approved', 'Rejected'):
        return jsonify({'error': 'Invalid status'}), 400
    subs = _load_submissions()
    rec = next((s for s in subs if s['id'] == sid), None)
    if not rec:
        return jsonify({'error': 'Not found'}), 404
    rec['status'] = new_status
    _save_submissions(subs)
    return jsonify({'ok': True})


@app.route('/api/next-claim-no', methods=['GET'])
@login_required
def next_claim_no():
    s = _load_settings()
    return jsonify({'next': s.get('claim_no_next', 1)})


@app.route('/api/settings', methods=['GET'])
@admin_required
def get_settings():
    return jsonify(_load_settings())


@app.route('/api/settings', methods=['PATCH'])
@admin_required
def update_settings():
    data = request.get_json(force=True)
    s = _load_settings()
    if 'claim_no_next' in data:
        try:
            val = int(data['claim_no_next'])
            if val < 1:
                return jsonify({'error': 'Must be >= 1'}), 400
            s['claim_no_next'] = val
        except (ValueError, TypeError):
            return jsonify({'error': 'Invalid number'}), 400
    _save_settings(s)
    return jsonify(s)


@app.route('/api/drafts', methods=['GET'])
@login_required
def list_drafts():
    drafts = _load_drafts()
    user_drafts = [d for d in drafts if d.get('username') == session['username']]
    user_drafts.sort(key=lambda d: d.get('updated_at', ''), reverse=True)
    return jsonify(user_drafts)


@app.route('/api/drafts/<did>', methods=['PUT'])
@login_required
def save_draft(did):
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400
    drafts = _load_drafts()
    rec = next((d for d in drafts if d['id'] == did), None)
    now = datetime.now().isoformat(timespec='seconds')
    if rec:
        if rec.get('username') != session['username']:
            return jsonify({'error': 'Forbidden'}), 403
        rec.update({**data, 'id': did, 'username': session['username'], 'updated_at': now})
    else:
        drafts.append({**data, 'id': did, 'username': session['username'], 'updated_at': now})
    _save_drafts(drafts)
    return jsonify({'id': did})


@app.route('/api/drafts/<did>', methods=['DELETE'])
@login_required
def delete_draft(did):
    drafts = _load_drafts()
    rec = next((d for d in drafts if d['id'] == did), None)
    if not rec:
        return jsonify({'error': 'Not found'}), 404
    if rec.get('username') != session['username']:
        return jsonify({'error': 'Forbidden'}), 403
    _save_drafts([d for d in drafts if d['id'] != did])
    return jsonify({'ok': True})


@app.route('/api/generate-excel', methods=['POST'])
@login_required
def generate_excel():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400

    wb = _build_workbook(data)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    employee_name = data.get('employee_name', 'Claim').replace(' ', '_').replace(',', '')
    claim_no = data.get('claim_no', '')
    filename = f"{employee_name}_Claim_{claim_no}.xlsx" if claim_no else f"{employee_name}_Claim.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def _build_workbook(data):
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()
    ws = wb.active
    ws.title = "Claim"

    ws.page_setup.paperSize  = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.scale      = 74
    ws.page_margins.left     = 0.45
    ws.page_margins.right    = 0.45
    ws.page_margins.top      = 0.5
    ws.page_margins.bottom   = 0.25
    ws.page_margins.header   = 0.3
    ws.page_margins.footer   = 0.3
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.column_dimensions['A'].width = 17.58
    ws.column_dimensions['B'].width = 44.25
    ws.column_dimensions['C'].width = 25.83
    ws.column_dimensions['D'].width = 14.33
    ws.column_dimensions['E'].width = 14.33

    CG = 'Century Gothic'
    fn  = Font(name=CG, size=11)
    fb  = Font(name=CG, size=11, bold=True)
    fs  = Font(name=CG, size=8)

    thin   = Side(style='thin')
    double = Side(style='double')
    b_all  = Border(left=thin, right=thin, top=thin, bottom=thin)
    b_tb   = Border(top=thin, bottom=thin)
    b_t    = Border(top=thin)
    b_td   = Border(top=thin, bottom=double)

    DATE_FMT = 'dd/mm/yyyy'

    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo.jpg')
    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width  = 135
        img.height = 21
        ws.add_image(img, 'A2')

    for r in range(1, 7):
        ws.row_dimensions[r].height = 18

    ws['D7'] = 'CLAIM PERIOD'
    ws['D7'].font = fb
    ws['D7'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('D7:E7')
    ws.row_dimensions[7].height = 18

    ws['A8'] = "Employee's Name:"
    ws['A8'].font = fn
    ws['A8'].alignment = Alignment(vertical='center')
    ws['B8'] = data.get('employee_name', '')
    ws['B8'].font = fn
    ws['B8'].alignment = Alignment(vertical='center', indent=1)
    ws['B8'].border = Border(bottom=thin)
    ws['D8'] = 'FROM'
    ws['D8'].font = fb
    ws['D8'].alignment = Alignment(horizontal='center', vertical='center')
    ws['E8'] = 'TO'
    ws['E8'].font = fb
    ws['E8'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[8].height = 18

    ws['A9'] = "Claim Form no.:"
    ws['A9'].font = fn
    ws['A9'].alignment = Alignment(vertical='center')
    ws['B9'] = data.get('claim_no', '')
    ws['B9'].font = fn
    ws['B9'].alignment = Alignment(vertical='center', indent=1)
    ws['B9'].border = b_tb

    period_from = data.get('period_from', '')
    period_to   = data.get('period_to', '')
    for col, val in [('D', period_from), ('E', period_to)]:
        cell = ws[f'{col}9']
        if val:
            try:
                cell.value = datetime.strptime(val, '%Y-%m-%d')
                cell.number_format = DATE_FMT
            except ValueError:
                cell.value = val
        cell.font      = fn
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = b_all
    ws.row_dimensions[9].height = 18

    ws.row_dimensions[10].height = 18
    ws.row_dimensions[11].height = 18

    ws['A12'] = 'DATE'
    ws['A12'].font = fb
    ws['A12'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A12'].border = b_all

    ws['B12'] = 'DESCRIPTION'
    ws['B12'].font = fb
    ws['B12'].alignment = Alignment(vertical='center')
    ws['B12'].border = b_all

    ws['C12'] = 'GST amount on each bill'
    ws['C12'].font = fb
    ws['C12'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['C12'].border = b_all

    currency = data.get('currency', 'SGD')
    ws['D12'] = f'TOTAL ({currency})'
    ws['D12'].font = fb
    ws['D12'].alignment = Alignment(horizontal='center', vertical='center')
    ws['D12'].border = b_all
    ws.merge_cells('D12:E12')
    ws.row_dimensions[12].height = 18

    items = data.get('items', [])
    for i in range(15):
        row  = 13 + i
        item = items[i] if i < len(items) else {}
        ws.row_dimensions[row].height = 20.15

        date_val = item.get('date', '')
        desc     = item.get('description', '')
        gst      = item.get('gst', '')
        total    = item.get('total', '')

        if date_val:
            try:
                ws[f'A{row}'] = datetime.strptime(date_val, '%Y-%m-%d')
                ws[f'A{row}'].number_format = DATE_FMT
            except ValueError:
                ws[f'A{row}'] = date_val
        ws[f'A{row}'].font = fn
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{row}'].border = b_all

        ws[f'B{row}'] = desc
        ws[f'B{row}'].font = fn
        ws[f'B{row}'].alignment = Alignment(vertical='center', wrap_text=True)
        ws[f'B{row}'].border = b_all

        if gst not in ('', None):
            try:
                ws[f'C{row}'] = float(gst)
                ws[f'C{row}'].number_format = '#,##0.00'
            except (ValueError, TypeError):
                ws[f'C{row}'] = gst
        ws[f'C{row}'].font = fn
        ws[f'C{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'C{row}'].border = b_all

        if total not in ('', None):
            try:
                ws[f'D{row}'] = float(total)
                ws[f'D{row}'].number_format = '#,##0.00'
            except (ValueError, TypeError):
                ws[f'D{row}'] = total
        ws[f'D{row}'].font = fn
        ws[f'D{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'D{row}'].border = b_all
        ws.merge_cells(f'D{row}:E{row}')

    ws.row_dimensions[28].height = 20.15
    ws['C28'] = 'Total Reimbursement'
    ws['C28'].font = fb
    ws['C28'].alignment = Alignment(horizontal='right', vertical='center')
    ws['C28'].border = b_td

    ws['D28'] = '=SUM(D13:D27)'
    ws['D28'].number_format = '#,##0.00'
    ws['D28'].font = fb
    ws['D28'].alignment = Alignment(horizontal='right', vertical='center')
    ws['D28'].border = b_td
    ws.merge_cells('D28:E28')

    for r in range(29, 38):
        ws.row_dimensions[r].height = 18

    ws.row_dimensions[38].height = 18
    for col, label in [('A', 'Received by'), ('B', 'Date'), ('C', 'Approved by'), ('D', 'Date')]:
        ws[f'{col}38'] = label
        ws[f'{col}38'].font = fn
        ws[f'{col}38'].alignment = Alignment(horizontal='center' if col != 'A' else 'left', vertical='center')
        ws[f'{col}38'].border = b_t

    for r in range(39, 42):
        ws.row_dimensions[r].height = 18

    ws.row_dimensions[42].height = 18
    ws['C42'] = 'Note:'
    ws['C42'].font = fb
    ws['C42'].alignment = Alignment(vertical='center')
    ws.merge_cells('C42:E42')

    note_text = data.get('notes', '')
    if note_text:
        ws['C43'] = note_text
        ws['C43'].font = fn
        ws['C43'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('C43:E46')

    for r in range(47, 54):
        ws.row_dimensions[r].height = 18

    ws.row_dimensions[54].height = 18
    ws['A54'] = COMPANY_NAME
    ws['A54'].font = fb
    ws['A54'].alignment = Alignment(vertical='center')
    ws.merge_cells('A54:E54')

    ws.row_dimensions[55].height = 18
    ws['A55'] = COMPANY_UEN
    ws['A55'].font = fs
    ws['A55'].alignment = Alignment(vertical='center')
    ws.merge_cells('A55:E55')

    ws.row_dimensions[56].height = 18
    ws['A56'] = COMPANY_ADDRESS
    ws['A56'].font = fs
    ws['A56'].alignment = Alignment(vertical='center')
    ws.merge_cells('A56:E56')

    attachments = data.get('attachments', [])
    if attachments:
        ws2 = wb.create_sheet("Attachments")
        for col, hdr in [('A', 'Item #'), ('B', 'Description'), ('C', 'Filename')]:
            ws2[f'{col}1'] = hdr
            ws2[f'{col}1'].font = Font(name=CG, bold=True, size=11)
        ws2.column_dimensions['A'].width = 10
        ws2.column_dimensions['B'].width = 50
        ws2.column_dimensions['C'].width = 50
        for i, att in enumerate(attachments, start=2):
            ws2[f'A{i}'] = att.get('item_index', '')
            ws2[f'B{i}'] = att.get('description', '')
            ws2[f'C{i}'] = att.get('original_name', '')
            for col in ('A', 'B', 'C'):
                ws2[f'{col}{i}'].font = Font(name=CG, size=11)

    return wb


if __name__ == '__main__':
    app.run(debug=True, port=5050)
